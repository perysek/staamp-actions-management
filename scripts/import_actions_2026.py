"""
One-off importer: management-review-actions-2026.xlsx → items (item_type='action').
Run from project root: .venv\\Scripts\\python.exe scripts/import_actions_2026.py

- Idempotent: skips an action whose (trimmed) title already exists.
- Topic-Area + COMMENTS are preserved inside the item description (no structured columns).
- No responsibles are assigned here (done later via the UI).
"""
import sys
import os
from datetime import date, datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from database.db import initialize_database, get_connection
from repositories.items.item_repository import ItemRepository

WORKBOOK = "management-review-actions-2026.xlsx"
SHEET = "Arkusz1"
TITLE_MAX = 140


def _to_date_str(value):
    """Excel cell → 'YYYY-MM-DD' or None. Handles datetime/date and Excel serial numbers."""
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        # Excel serial date (1900 date system; 1899-12-30 epoch)
        return (date(1899, 12, 30) + timedelta(days=int(value))).strftime('%Y-%m-%d')
    return None


def _status_from_pdca(pdca, completed_at):
    """PDCA letter → item status. Completed_at present overrides to 'done'."""
    if completed_at not in (None, ""):
        return 'done'
    letter = (str(pdca).strip().upper()[:1] if pdca not in (None, "") else "")
    if letter == 'A':
        return 'done'
    if letter in ('C', 'D'):
        return 'in_progress'
    return 'open'  # 'P' or blank


def main():
    initialize_database()
    item_repo = ItemRepository()

    # created_by — use an existing superuser if present (nullable otherwise)
    with get_connection() as conn:
        row = conn.execute("SELECT id FROM users WHERE role = 'superuser' ORDER BY id LIMIT 1").fetchone()
        created_by = row['id'] if row else None
        existing_titles = {r['title'] for r in conn.execute("SELECT title FROM items").fetchall()}
    if created_by is None:
        print("UWAGA: brak konta superuser — created_by ustawione na NULL. "
              "Uruchom najpierw scripts/seed_users.py, aby przypisać autora.")

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]

    imported = skipped = 0
    for r in range(2, ws.max_row + 1):
        topic_area = ws.cell(row=r, column=2).value
        action = ws.cell(row=r, column=3).value
        scheduled_at = ws.cell(row=r, column=5).value
        completed_at = ws.cell(row=r, column=6).value
        comments = ws.cell(row=r, column=7).value
        pdca = ws.cell(row=r, column=8).value

        if action in (None, ""):
            continue  # skip rows without an Action

        action_full = str(action).strip()
        title = action_full[:TITLE_MAX]
        if title in existing_titles:
            skipped += 1
            continue

        description = (
            f"[Obszar: {str(topic_area).strip() if topic_area else '—'}]\n\n"
            f"{action_full}\n\n"
            f"Uwagi: {str(comments).strip() if comments else '—'}"
        )
        due_date = _to_date_str(scheduled_at)
        status = _status_from_pdca(pdca, completed_at)

        item_repo.create('action', title, description, status, due_date, created_by)
        existing_titles.add(title)
        imported += 1

    print(f"Imported {imported} actions" + (f" (skipped {skipped} existing)" if skipped else ""))


if __name__ == "__main__":
    main()
